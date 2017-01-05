import unicodedata
from django.core.exceptions import ObjectDoesNotExist
from django.utils.translation import ugettext as _

from django.contrib import messages

from django.contrib.auth.models import User
import courses.models as models

import logging

from django.http.response import HttpResponse
from datetime import date

log = logging.getLogger('tq')

from .emailcenter import *

import re


# Create your services here.
def get_all_offerings():
    return models.Offering.objects.order_by('period__date_from', '-active')


def get_offerings_to_display(request=None, force_preview=False):
    # return offerings that have display flag on and order them with increasing start
    if force_preview or (request and request.user.is_staff):
        return models.Offering.objects.filter(Q(display=True) | Q(period__date_to__gte=date.today())).order_by(
            'period__date_from')
    else:
        return models.Offering.objects.filter(display=True).order_by('-active', 'period__date_from')


def get_current_active_offering():
    return models.Offering.objects.filter(active=True).order_by('period__date_from').first()


def get_subsequent_offering():
    res = models.Offering.objects.filter(period__date_from__gte=date.today()).order_by(
        'period__date_from').all()
    if len(res) > 0:
        return res[0]
    else:
        return None


def get_or_create_user(user_data):
    user = find_user(user_data)
    if user:
        return update_user(user, user_data)
    else:
        return create_user(user_data)


def update_user(user, user_data):
    if 'email' in user_data:
        user.email = user_data['email']
    if 'first_name' in user_data:
        user.first_name = user_data['first_name']
    if 'last_name' in user_data:
        user.last_name = user_data['last_name']
    user.save()

    userprofile = get_or_create_userprofile(user)

    # convenience method. if key is not given, assume same as attr
    def set_if_given(attr, key=None):
        if not key:
            key = attr
        if key in user_data:
            setattr(userprofile, attr, user_data[key])

    set_if_given('legi')
    set_if_given('gender')
    set_if_given('phone_number')
    set_if_given('student_status')
    set_if_given('body_height')
    set_if_given('newsletter')
    set_if_given('get_involved')

    if all((key in user_data) for key in ['street', 'plz', 'city']):
        userprofile.address = models.Address.objects.create_from_user_data(user_data)

    userprofile.save()

    return user


def find_user(user_data):
    # check if user already exists
    fn = user_data['first_name']
    ln = user_data['last_name']

    q_email = Q(email=user_data['email'])
    q_address = Q(profile__address__plz=user_data['plz']) & Q(profile__address__street=user_data['street']) & Q(
        profile__address__city=user_data['city'])
    q = q_email | q_address
    qs = User.objects.filter(first_name=fn, last_name=ln).filter(q).all()
    if qs.count() > 0:
        return qs[0]
    else:
        return None


def create_user(user_data):
    fn = user_data['first_name']
    ln = user_data['last_name']

    user = User.objects.create_user(generate_username(fn, ln), email=user_data['email'],
                                    password=User.objects.make_random_password())

    update_user(user, user_data)
    return user


def generate_username(first_name, last_name):
    username = "{}_{}".format(first_name, last_name)
    return find_unused_username_variant(clean_username(username)).lower()


def find_unused_username_variant(name, ignore=None):
    un = name
    i = 1
    while User.objects.exclude(username=ignore).filter(username=un).count() > 0:
        un = name + str(i)
        i += 1
    return un


def clean_username(name):
    """first try to find ascii similar character, then strip away disallowed characters still left"""
    name = unicodedata.normalize('NFKD', name)
    return re.sub('[^0-9a-zA-Z+-.@_]+', '', name)


def subscribe(course_id, user1_data, user2_data=None):
    res = dict()

    course = models.Course.objects.get(id=course_id)
    user1 = get_or_create_user(user1_data)
    if user2_data:
        user2 = get_or_create_user(user2_data)
    else:
        user2 = None

    if user1 == user2:
        res['tag'] = 'danger'
        res['text'] = 'Du kannst dich nicht mit dir selbst anmelden!'
    elif models.Subscribe.objects.filter(user=user1, course__id=course_id).count() > 0:
        res['tag'] = 'danger'
        res['text'] = 'Du ({}) bist schon für diesen Kurs angemeldet!'.format(user1.first_name)
        res['long_text'] = 'Wenn du ein Fehler bei der Anmeldung gemacht hast, wende dich an tanzen@tq.vseth.ch'
    elif user2 is not None and models.Subscribe.objects.filter(user=user2, course__id=course_id).count() > 0:
        res['tag'] = 'danger'
        res['text'] = 'Dein Partner {} ist schon für diesen Kurs angemeldet!'.format(user2.first_name)
        res['long_text'] = 'Wenn du ein Fehler bei der Anmeldung gemacht hast, wende dich an tanzen@tq.vseth.ch'
    else:
        subscription = models.Subscribe(user=user1, course=course, partner=user2,
                                        experience=user1_data['experience'],
                                        comment=user1_data['comment'])
        subscription.derive_matching_state()
        subscription.save()
        send_subscription_confirmation(subscription)

        if user2:
            subscription.matching_state = models.Subscribe.MatchingState.COUPLE
            subscription.save()

            subscription2 = models.Subscribe(user=user2, course=course, partner=user1,
                                             experience=user2_data['experience'], comment=user2_data['comment'])
            subscription2.matching_state = models.Subscribe.MatchingState.COUPLE
            subscription2.save()
            send_subscription_confirmation(subscription2)

        res['tag'] = 'info'
        res['text'] = 'Anmeldung erfolgreich.'
        res['long_text'] = 'Du erhältst in Kürze eine Emailbestätigung.'

    return res


# creates a copy of course and sets its offering to the next offering in the future
def copy_course(course, to=None):
    if to is None:
        to = get_subsequent_offering()
    if to is not None:
        course = course.copy()
        course.offering = to
        course.active = False
        course.save()


# matches partners within the same course, considering their subscription time (fairness!) and respects also body_height (second criteria)
DEFAULT_BODY_HEIGHT = 170


def match_partners(subscriptions, request=None):
    courses = subscriptions.values_list('course', flat=True)
    match_count = 0
    for course_id in courses:
        single = subscriptions.filter(course__id=course_id, partner__isnull=True).all()
        sm = single.filter(user__profile__gender=models.UserProfile.Gender.MEN).order_by('date').all()
        sw = single.filter(user__profile__gender=models.UserProfile.Gender.WOMAN).order_by('date').all()
        c = min(sm.count(), sw.count())
        sm = list(sm[0:c])  # list() enforces evaluation of queryset
        sw = list(sw[0:c])
        sm.sort(key=lambda
            x: x.user.profile.body_height if x.user.profile and x.user.profile.body_height else DEFAULT_BODY_HEIGHT)
        sw.sort(key=lambda
            x: x.user.profile.body_height if x.user.profile and x.user.profile.body_height else DEFAULT_BODY_HEIGHT)
        while c > 0:
            c = c - 1
            m = sm[c]
            w = sw[c]
            m.partner = w.user
            m.matching_state = models.Subscribe.MatchingState.MATCHED
            m.save()
            w.partner = m.user
            w.matching_state = models.Subscribe.MatchingState.MATCHED
            w.save()
            match_count += 1
    if match_count:
        messages.add_message(request, messages.SUCCESS,
                             _(u'{} couples matched successfully').format(match_count))


def unmatch_partners(subscriptions):
    for s in subscriptions.all():
        if s.state == models.Subscribe.State.NEW:
            allowed_states = [models.Subscribe.MatchingState.MATCHED, models.Subscribe.MatchingState.COUPLE]
            partner_subs = subscriptions.filter(user=s.partner, course=s.course)
            if partner_subs.count() == 1 and s.matching_state in allowed_states and partner_subs.first().matching_state in allowed_states:
                _unmatch_person(s)
                _unmatch_person(partner_subs.first())


def _unmatch_person(subscription):
    subscription.partner = None
    subscription.matching_state = models.Subscribe.MatchingState.TO_REMATCH
    subscription.save()


class CourseException(Exception):
    def __init__(self, value):
        self.value = value

    def __str__(self):
        return self.value


class NoPartnerException(CourseException):
    def __init__(self):
        super(NoPartnerException, self).__init__(_('This subscription has no partner set'))


# sends a confirmation mail if subscription is confirmed (by some other method) and no confirmation mail was sent before
def confirm_subscription(subscription, request=None, allow_single_in_couple_course=False):
    # check: only people with partner are confirmed (in couple courses)
    if not allow_single_in_couple_course and subscription.course.type.couple_course and subscription.partner is None:
        raise NoPartnerException()

    if subscription.state == models.Subscribe.State.NEW:
        subscription.state = models.Subscribe.State.CONFIRMED
        subscription.save()

        m = send_participation_confirmation(subscription)
        if m:
            # log that we sent the confirmation
            c = models.Confirmation(subscription=subscription, mail=m)
            c.save()
            return True
        else:
            return False
    else:
        return False


# same as confirm_subscription, but for multiple subscriptions at once
MESSAGE_NO_PARTNER_SET = _(u'{} subscriptions were not confirmed because no partner set')


def confirm_subscriptions(subscriptions, request=None, allow_single_in_couple_course=False):
    no_partner_count = 0
    confirmed_count = 0
    for subscription in subscriptions:
        try:
            if confirm_subscription(subscription, request, allow_single_in_couple_course):
                confirmed_count += 1
        except NoPartnerException as e:
            no_partner_count += 1

    if no_partner_count:  # if any subscriptions not confirmed due to missing partner
        log.warning(MESSAGE_NO_PARTNER_SET.format(no_partner_count))
        if request:
            messages.add_message(request, messages.WARNING, MESSAGE_NO_PARTNER_SET.format(no_partner_count))
    if confirmed_count:
        messages.add_message(request, messages.SUCCESS,
                             _(u'{} of {} confirmed successfully').format(confirmed_count, len(subscriptions)))


def unconfirm_subscriptions(subscriptions, request=None):
    for s in subscriptions.all():
        if s.state == models.Subscribe.State.CONFIRMED:
            s.state = models.Subscribe.State.NEW
            s.save()


# sends a rejection mail if subscription is rejected (by some other method) and no rejection mail was sent before
def reject_subscription(subscription, reason=None, send_email=True):
    subscription.state = models.Subscribe.State.REJECTED
    subscription.save()
    if not reason:
        reason = detect_rejection_reason(subscription)
    c = models.Rejection(subscription=subscription, reason=reason, mail_sent=False)
    c.save()

    if send_email and models.Rejection.objects.filter(subscription=subscription, mail_sent=True).count() == 0:
        # if ensures that no mail was ever sent due to a rejection to this user

        # save if we sent the mail
        c.mail = send_rejection(subscription, reason)
        c.mail_sent = c.mail is not None
        c.save()


# same as reject_subscription, but for multiple subscriptions at once
def reject_subscriptions(subscriptions, reason=None, send_email=True):
    for subscription in subscriptions:
        reject_subscription(subscription, reason, send_email)


def unreject_subscriptions(subscriptions, request=None):
    unrejected_count = 0
    for subscription in subscriptions:
        if subscription.state == models.Subscribe.State.REJECTED:
            subscription.state = models.Subscribe.State.NEW
            subscription.save()
            unrejected_count += 1
    if unrejected_count:
        messages.add_message(request, messages.SUCCESS,
                             _(u'{} unrejected successfully').format(unrejected_count))


def welcome_teacher(teach):
    if not teach.welcomed:
        teach.welcomed = True
        teach.save()

        m = send_teacher_welcome(teach)
        if m:
            # log that we sent the confirmation
            c = models.TeacherWelcome(teach=teach, mail=m)
            c.save()
            return True
        else:
            return False
    else:
        return False


def welcome_teachers(courses, request):
    count = 0
    total = 0
    for course in courses:
        for teach in course.teaching.all():
            total += 1
            if welcome_teacher(teach):
                count += 1
    messages.add_message(request, messages.SUCCESS,
                         _(u'{} of {} welcomed successfully').format(count, total))


def welcome_teachers_reset_flag(courses, request):
    count = 0
    total = 0
    for course in courses:
        for teach in course.teaching.all():
            if teach.welcomed:
                count += 1
                teach.welcomed = False
                teach.save()
            total += 1
    messages.add_message(request, messages.SUCCESS,
                         _(u'{} of {} teachers reset successfully').format(count, total))


def get_or_create_userprofile(user):
    try:
        return models.UserProfile.objects.get(user=user)
    except ObjectDoesNotExist:
        userprofile = models.UserProfile(user=user)
        return userprofile


# finds a list of courses the 'user' did already and that are somehow relevant for 'course'
def calculate_relevant_experience(user, course):
    relevant_exp = [style.id for style in course.type.styles.all()]
    return [s.course for s in
            models.Subscribe.objects.filter(user=user, state__in=[models.Subscribe.State.CONFIRMED,
                                                                  models.Subscribe.State.PAYED,
                                                                  models.Subscribe.State.COMPLETED],
                                            course__type__styles__id__in=relevant_exp).exclude(
                course=course).order_by('course__type__level').distinct().all()]


def format_prices(price_with_legi, price_without_legi, price_special=None):
    if price_special:
        return price_special
    elif price_with_legi and price_without_legi:
        if price_with_legi == price_without_legi:
            r = "{} CHF".format(price_with_legi.__str__())
        else:
            r = "mit Legi {} / sonst {} CHF".format(price_with_legi.__str__(), price_without_legi.__str__())
    elif price_without_legi:
        r = "mit Legi freier Eintritt (sonst {} CHF)".format(price_without_legi.__str__())
    else:
        r = None  # handle this case in template!
    return r


import zipfile
import unicodecsv
from io import BytesIO

import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

INVALID_TITLE_CHARS = re.compile(r'[^\w\-_ ]', re.IGNORECASE | re.UNICODE)


# exports the subscriptions of course with course_id to fileobj (e.g. a HttpResponse)
def export_subscriptions(course_ids, export_format):
    def create_xlsx_sheet(wb, course_id, course_name):
        # strip away unallowed characters and restrict to 30 characters
        ws = wb.create_sheet(title=INVALID_TITLE_CHARS.sub("", course_name)[:30])

        row_num = 0

        columns = [
            (u"Vorname", 20),
            (u"Nachname", 20),
            (u"Geschlecht", 10),
            (u"E-Mail", 50),
            (u"Mobile", 30),
            (u"Legi-Nr.", 30),
            (u"Preis", 10),
            (u"Noch zu bezahlen", 10),
            (u"Erfahrung", 60),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            font = c.font.copy()
            font.bold = True
            c.font = font
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for s in models.Subscribe.objects.accepted().filter(course__id=course_id).order_by('user__first_name'):
            row_num += 1
            row = [s.user.first_name, s.user.last_name, s.user.profile.gender, s.user.email,
                   s.user.profile.phone_number, s.user.profile.legi, s.get_price_to_pay(),
                   0 if s.payed() else s.get_price_to_pay(), s.experience]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

                alignment = c.alignment.copy()
                alignment.wrap_text = True
                c.alignment = alignment

    if len(course_ids) == 1:
        course_id = course_ids[0]
        course_name = models.Course.objects.get(id=course_id).name
        filename = 'Kursteilnehmer-{}'.format(course_name)
        if export_format == 'csv':
            # Create the HttpResponse object with the appropriate CSV header.
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(filename)

            writer = unicodecsv.writer(response)

            writer.writerow([u'Vorname', 'Nachname', 'Geschlecht', 'E-Mail', 'Mobile', 'Legi-Nr.', 'Zu bezahlen',
                             'Erfahrung'])
            for s in models.Subscribe.objects.accepted().filter(course__id=course_id).order_by(
                    'user__first_name'):
                row = [s.user.first_name, s.user.last_name, s.user.profile.gender, s.user.email,
                       s.user.profile.phone_number, s.user.profile.legi, s.get_price_to_pay(), s.experience]
                writer.writerow(row)

            return response
        if export_format == 'csv_google':
            # Create the HttpResponse object with the appropriate CSV header.
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(filename)

            writer = unicodecsv.writer(response)

            writer.writerow(
                [u'Given Name', 'Family Name', 'Gender', 'E-mail 1 - Type', 'E-mail 1 - Value', 'Phone 1 - Type',
                 'Phone 1 - Value'])
            for s in models.Subscribe.objects.accepted().filter(course__id=course_id).order_by(
                    'user__first_name'):
                row = [s.user.first_name, s.user.last_name, s.user.profile.gender, '* Private', s.user.email,
                       '* Private',
                       s.user.profile.phone_number]
                writer.writerow(row)

            return response
        elif export_format == 'xlsx':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(filename)
            wb = openpyxl.Workbook()
            # remove preinitialized sheet
            wb.remove_sheet(wb.get_active_sheet())

            create_xlsx_sheet(wb, course_id, course_name)

            wb.save(response)
            return response
        else:
            return None
    elif len(course_ids) > 1:
        if export_format == 'csv':
            zipped_file = BytesIO()  # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes
            with zipfile.ZipFile(zipped_file, 'w') as f:
                for course_id in course_ids:
                    fileobj = BytesIO()  # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes
                    writer = unicodecsv.writer(fileobj, encoding='utf-8')

                    writer.writerow(
                        ['Vorname', 'Nachname', 'Geschlecht', 'E-Mail', 'Mobile', 'Legi-Nr.', 'Zu bezahlen',
                         'Erfahrung'])
                    for s in models.Subscribe.objects.accepted().filter(course__id=course_id).order_by(
                            'user__first_name'):
                        l = [s.user.first_name, s.user.last_name, s.user.profile.gender, s.user.email,
                             s.user.profile.phone_number, s.user.profile.legi, s.get_price_to_pay(), s.experience]
                        writer.writerow(l)
                    f.writestr(u'Kursteilnehmer/{}.csv'.format(models.Course.objects.get(id=course_id).name),
                               fileobj.getvalue())
                    fileobj.seek(0)

            zipped_file.seek(0)
            response = HttpResponse(zipped_file, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=Kursteilnehmer.zip'
            response['Content-Length'] = zipped_file.tell()

            return response
        if export_format == 'csv_google':
            zipped_file = BytesIO()  # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes
            with zipfile.ZipFile(zipped_file, 'w') as f:
                for course_id in course_ids:
                    fileobj = BytesIO()  # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes
                    writer = unicodecsv.writer(fileobj, encoding='utf-8')

                    writer.writerow(
                        [u'Given Name', 'Family Name', 'Gender', 'E-mail 1 - Type', 'E-mail 1 - Value',
                         'Phone 1 - Type',
                         'Phone 1 - Value'])
                    for s in models.Subscribe.objects.accepted().filter(course__id=course_id).order_by(
                            'user__first_name'):
                        row = [s.user.first_name, s.user.last_name, s.user.profile.gender, '* Private', s.user.email,
                               '* Private',
                               s.user.profile.phone_number]
                        writer.writerow(row)
                    f.writestr(u'Kursteilnehmer/{}.csv'.format(models.Course.objects.get(id=course_id).name),
                               fileobj.getvalue())
                    fileobj.seek(0)

            zipped_file.seek(0)
            response = HttpResponse(zipped_file, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=Kursteilnehmer.zip'
            response['Content-Length'] = zipped_file.tell()

            return response
        elif export_format == 'xlsx':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Kursteilnehmer.xlsx'
            wb = openpyxl.Workbook()
            # remove preinitialized sheet
            wb.remove_sheet(wb.get_active_sheet())

            for course_id in course_ids:
                create_xlsx_sheet(wb, course_id, models.Course.objects.get(id=course_id).name)

            wb.save(response)
            return response
        else:
            return None
    else:
        return None


from django.db import transaction
from django.apps import apps
from django.db.models import Model
from django.contrib.contenttypes.fields import GenericForeignKey
from django.db.models import Q

from django.contrib.auth.models import User


# find duplicates for all users in the system
def find_duplicate_users():
    ret = dict()
    done = set()
    for user in User.objects.all():
        d = find_user_duplicates_ids(user)
        if len(d) > 1 and user.id not in done:
            ret[user.id] = d[1:]
            done.update(d)
    return ret


# finds duplicates of the given user
def find_user_duplicates_ids(user):
    """
    Finds duplicates of user. Returns a list of ids including user's id, ordered by the date the user joined the system.
    """
    candidates = list(User.objects.filter(first_name=user.first_name, last_name=user.last_name).order_by('date_joined'))
    # candidates intentionally also includes the passed user
    ret = set()
    for c in candidates:
        ret.add(c.id)
    return list(ret)


def merge_duplicate_users(to_merge):
    for primary, aliases in to_merge.items():
        merge_model_objects(primary, aliases, False)


def merge_duplicate_users_by_ids(to_merge):
    for primary, aliases in to_merge.items():
        merge_model_objects(User.objects.get(id=primary), list(User.objects.filter(id__in=aliases)), False)


# Based on https://gist.github.com/NicholasMerrill/7c395aa3634b2f2a0cb4
def merge_model_objects(primary_object, alias_objects=None, keep_old=False):
    """
    Use this function to merge model objects (i.e. Users, Organizations, Polls,
    etc.) and migrate all of the related fields from the alias objects to the
    primary object.

    Usage:
    from django.contrib.auth.models import User
    primary_user = User.objects.get(email='good_email@example.com')
    duplicate_user = User.objects.get(email='good_email+duplicate@example.com')
    merge_model_objects(primary_user, duplicate_user)
    """
    if not alias_objects: alias_objects = []
    if not isinstance(alias_objects, list):
        alias_objects = [alias_objects]

    # check that all aliases are the same class as primary one and that
    # they are subclass of model
    primary_class = primary_object.__class__

    if not issubclass(primary_class, Model):
        raise TypeError('Only django.db.models.Model subclasses can be merged')

    for alias_object in alias_objects:
        if not isinstance(alias_object, primary_class):
            raise TypeError('Only models of same class can be merged')

    # Get a list of all GenericForeignKeys in all models
    # TODO: this is a bit of a hack, since the generics framework should provide a similar
    # method to the ForeignKey field for accessing the generic related fields.
    generic_fields = []
    for model in apps.get_models():
        for field_name, field in filter(lambda x: isinstance(x[1], GenericForeignKey), model.__dict__.items()):
            generic_fields.append(field)

    blank_local_fields = set([field.attname for field in primary_object._meta.local_fields if
                              getattr(primary_object, field.attname) in [None, '']])

    # Loop through all alias objects and migrate their data to the primary object.
    for alias_object in alias_objects:
        # Migrate all foreign key references from alias object to primary object.
        for related_object in alias_object._meta.get_all_related_objects():
            log.info("related object: {}".format(related_object))
            # The variable name on the alias_object model.
            alias_varname = related_object.get_accessor_name()
            # The variable name on the related model.
            obj_varname = related_object.field.name
            if hasattr(alias_object, alias_varname):
                related_objects = getattr(alias_object, alias_varname)
                if hasattr(related_objects, 'all'):
                    for obj in related_objects.all():
                        setattr(obj, obj_varname, primary_object)
                        obj.save()
                else:
                    # `related_objects` is a one-to-one field.
                    # Merge related one-to-one fields.
                    alias_related_object = related_objects
                    primary_related_object = getattr(primary_object, alias_varname)
                    # The delete will cascade later if `keep_old` is False.
                    # Otherwise, could violate a not-null one-to-one field constraint.
                    merge_model_objects(primary_related_object, alias_related_object, keep_old=True)
            else:
                # TODO This is kind of a hack: we just display a message and do not merge that failing relation!
                log.warning("some parts could not be merged: {} has no variable {} induced from {}".format(alias_object,
                                                                                                           alias_varname,
                                                                                                           related_object))

        # Migrate all many to many references from alias object to primary object.
        for related_many_object in alias_object._meta.get_all_related_many_to_many_objects():
            log.info("many to many: {}".format(related_many_object))
            alias_varname = related_many_object.get_accessor_name()
            obj_varname = related_many_object.field.name

            if alias_varname is not None:
                # standard case
                related_many_objects = getattr(alias_object, alias_varname).all()
            else:
                # special case, symmetrical relation, no reverse accessor
                related_many_objects = getattr(alias_object, obj_varname).all()
            for obj in related_many_objects.all():
                try:
                    getattr(obj, obj_varname).remove(alias_object)
                    getattr(obj, obj_varname).add(primary_object)
                except AttributeError:
                    pass  # TODO kind of a hack because we do not know how to check if m2m related object has a through model defined (in which case we can not use remove/add)

        # Migrate all generic foreign key references from alias object to primary object.
        for field in generic_fields:
            filter_kwargs = {}
            filter_kwargs[field.fk_field] = alias_object._get_pk_val()
            filter_kwargs[field.ct_field] = field.get_content_type(alias_object)
            for generic_related_object in field.model.objects.filter(**filter_kwargs):
                setattr(generic_related_object, field.name, primary_object)
                generic_related_object.save()

        # Try to fill all missing values in primary object by values of duplicates
        filled_up = set()
        for field_name in blank_local_fields:
            val = getattr(alias_object, field_name)
            if val not in [None, '']:
                setattr(primary_object, field_name, val)
                filled_up.add(field_name)
        blank_local_fields -= filled_up

        if not keep_old:
            alias_object.delete()
    primary_object.save()
    return primary_object
